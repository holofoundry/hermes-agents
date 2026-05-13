from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except Exception:  # pragma: no cover
    Workbook = None
    Font = None
    PatternFill = None

Number = float | int


def _num(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_div(numerator: Number, denominator: Number) -> Optional[float]:
    denominator = _num(denominator)
    if denominator == 0:
        return None
    return _num(numerator) / denominator


def _percent(value: Optional[float]) -> Optional[float]:
    return None if value is None else round(value * 100, 2)


def _stats(values: Iterable[Optional[float]]) -> Dict[str, Optional[float]]:
    clean = sorted(v for v in values if v is not None)
    if not clean:
        return {"count": 0, "min": None, "q1": None, "median": None, "mean": None, "q3": None, "max": None}
    def q(p: float) -> float:
        idx = (len(clean) - 1) * p
        lo = int(idx)
        hi = min(lo + 1, len(clean) - 1)
        frac = idx - lo
        return clean[lo] * (1 - frac) + clean[hi] * frac
    return {
        "count": len(clean),
        "min": round(clean[0], 2),
        "q1": round(q(0.25), 2),
        "median": round(median(clean), 2),
        "mean": round(mean(clean), 2),
        "q3": round(q(0.75), 2),
        "max": round(clean[-1], 2),
    }


def build_dcf_model(payload: Dict[str, Any], output_path: Optional[str] = None) -> Dict[str, Any]:
    """Build a simple formula-driven DCF workbook and return valuation summary.

    Expected payload keys:
    company, ticker, base_year_revenue, revenue_growth, ebit_margin, tax_rate,
    da_pct_revenue, capex_pct_revenue, nwc_pct_revenue, discount_rate,
    terminal_growth_rate, cash, debt, shares_outstanding, projection_years.
    """
    years = int(payload.get("projection_years", 5))
    years = max(1, min(years, 10))

    company = payload.get("company") or payload.get("ticker") or "Company"
    ticker = payload.get("ticker", "")
    revenue = _num(payload.get("base_year_revenue"), 0)
    growth = _num(payload.get("revenue_growth"), 0.05)
    ebit_margin = _num(payload.get("ebit_margin"), 0.18)
    tax_rate = _num(payload.get("tax_rate"), 0.24)
    da_pct = _num(payload.get("da_pct_revenue"), 0.03)
    capex_pct = _num(payload.get("capex_pct_revenue"), 0.04)
    nwc_pct = _num(payload.get("nwc_pct_revenue"), 0.01)
    discount_rate = _num(payload.get("discount_rate"), 0.10)
    terminal_growth = _num(payload.get("terminal_growth_rate"), 0.025)
    cash = _num(payload.get("cash"), 0)
    debt = _num(payload.get("debt"), 0)
    shares = _num(payload.get("shares_outstanding"), 1)

    missing = [k for k in ["base_year_revenue", "shares_outstanding"] if payload.get(k) in [None, ""]]
    if discount_rate <= terminal_growth:
        missing.append("discount_rate must be greater than terminal_growth_rate")

    rows: List[Dict[str, float]] = []
    prev_revenue = revenue
    pv_fcf_total = 0.0
    for i in range(1, years + 1):
        rev = prev_revenue * (1 + growth)
        ebit = rev * ebit_margin
        nopat = ebit * (1 - tax_rate)
        da = rev * da_pct
        capex = rev * capex_pct
        nwc = rev * nwc_pct
        fcf = nopat + da - capex - nwc
        pv = fcf / ((1 + discount_rate) ** i)
        pv_fcf_total += pv
        rows.append({
            "year": i,
            "revenue": rev,
            "ebit": ebit,
            "nopat": nopat,
            "da": da,
            "capex": capex,
            "nwc_investment": nwc,
            "free_cash_flow": fcf,
            "pv_fcf": pv,
        })
        prev_revenue = rev

    terminal_fcf = rows[-1]["free_cash_flow"] * (1 + terminal_growth)
    terminal_value = terminal_fcf / max(discount_rate - terminal_growth, 0.000001)
    pv_terminal_value = terminal_value / ((1 + discount_rate) ** years)
    enterprise_value = pv_fcf_total + pv_terminal_value
    equity_value = enterprise_value + cash - debt
    value_per_share = equity_value / shares if shares else None

    summary = {
        "company": company,
        "ticker": ticker,
        "currency": payload.get("currency", "local"),
        "projection_years": years,
        "enterprise_value": round(enterprise_value, 2),
        "equity_value": round(equity_value, 2),
        "value_per_share": None if value_per_share is None else round(value_per_share, 2),
        "pv_fcf": round(pv_fcf_total, 2),
        "pv_terminal_value": round(pv_terminal_value, 2),
        "terminal_value_pct_ev": _percent(_safe_div(pv_terminal_value, enterprise_value)),
        "missing_or_warnings": missing,
        "human_review_required": True,
    }

    if output_path:
        if Workbook is None:
            raise RuntimeError("openpyxl is required to write XLSX output")
        _write_dcf_workbook(payload, rows, summary, output_path)
        summary["output_path"] = str(output_path)

    return {"summary": summary, "projection": rows}


def _write_dcf_workbook(payload: Dict[str, Any], rows: List[Dict[str, float]], summary: Dict[str, Any], output_path: str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"] = "DCF Valuation Summary"
    ws["A1"].font = Font(bold=True, size=14)
    r = 3
    for k, v in summary.items():
        ws.cell(r, 1, k)
        ws.cell(r, 2, str(v) if isinstance(v, list) else v)
        ws.cell(r, 1).font = bold
        r += 1

    inputs = wb.create_sheet("Inputs")
    inputs.append(["Input", "Value"])
    for cell in inputs[1]:
        cell.font = bold
        cell.fill = fill
    for k, v in payload.items():
        inputs.append([k, v])

    proj = wb.create_sheet("Projection")
    headers = ["year", "revenue", "ebit", "nopat", "da", "capex", "nwc_investment", "free_cash_flow", "pv_fcf"]
    proj.append(headers)
    for cell in proj[1]:
        cell.font = bold
        cell.fill = fill
    for row in rows:
        proj.append([row[h] for h in headers])
    for sheet in wb.worksheets:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = max(14, min(28, max(len(str(c.value or "")) for c in col) + 2))
    wb.save(path)


def build_comps_summary(payload: Dict[str, Any]) -> Dict[str, Any]:
    peers = payload.get("peers", [])
    enriched = []
    for p in peers:
        ev = _num(p.get("enterprise_value"))
        market_cap = _num(p.get("market_cap"))
        revenue = _num(p.get("revenue"))
        ebitda = _num(p.get("ebitda"))
        ebit = _num(p.get("ebit"))
        net_income = _num(p.get("net_income"))
        enriched.append({
            **p,
            "ev_revenue": _safe_div(ev, revenue),
            "ev_ebitda": _safe_div(ev, ebitda),
            "ev_ebit": _safe_div(ev, ebit),
            "pe": _safe_div(market_cap, net_income),
            "ebitda_margin_pct": _percent(_safe_div(ebitda, revenue)),
        })
    return {
        "target": payload.get("target"),
        "as_of_date": payload.get("as_of_date"),
        "peer_count": len(enriched),
        "multiples": {
            "ev_revenue": _stats(p.get("ev_revenue") for p in enriched),
            "ev_ebitda": _stats(p.get("ev_ebitda") for p in enriched),
            "ev_ebit": _stats(p.get("ev_ebit") for p in enriched),
            "pe": _stats(p.get("pe") for p in enriched),
        },
        "peers": enriched,
        "data_quality_notes": _comps_quality_notes(enriched),
        "human_review_required": True,
    }


def _comps_quality_notes(peers: List[Dict[str, Any]]) -> List[str]:
    notes = []
    for p in peers:
        name = p.get("name") or p.get("ticker") or "Unnamed peer"
        for metric in ["ev_revenue", "ev_ebitda", "ev_ebit", "pe"]:
            if p.get(metric) is None:
                notes.append(f"{name}: {metric} unavailable because denominator is missing or zero.")
    return notes


def reconcile_ledgers(payload: Dict[str, Any]) -> Dict[str, Any]:
    gl = payload.get("gl", [])
    sub = payload.get("subledger", [])
    tolerance = _num(payload.get("tolerance"), 0.01)

    matched_sub = set()
    matches = []
    breaks = []

    for i, g in enumerate(gl):
        g_ref = str(g.get("reference", "")).strip().lower()
        g_amt = _num(g.get("amount"))
        g_ccy = g.get("currency")
        best: Optional[Tuple[int, Dict[str, Any]]] = None
        for j, s in enumerate(sub):
            if j in matched_sub:
                continue
            s_ref = str(s.get("reference", "")).strip().lower()
            s_amt = _num(s.get("amount"))
            s_ccy = s.get("currency")
            if g_ref and g_ref == s_ref and g_ccy == s_ccy:
                best = (j, s)
                break
            if abs(g_amt - s_amt) <= tolerance and g_ccy == s_ccy and g.get("date") == s.get("date"):
                best = (j, s)
        if best:
            j, s = best
            matched_sub.add(j)
            diff = round(g_amt - _num(s.get("amount")), 2)
            if abs(diff) <= tolerance:
                matches.append({"gl": g, "subledger": s, "difference": diff})
            else:
                breaks.append({"type": "amount_mismatch", "gl": g, "subledger": s, "difference": diff})
        else:
            breaks.append({"type": "missing_in_subledger", "gl": g})

    for j, s in enumerate(sub):
        if j not in matched_sub:
            breaks.append({"type": "missing_in_gl", "subledger": s})

    total_break = sum(_num(b.get("difference")) for b in breaks)
    return {
        "entity": payload.get("entity"),
        "period": payload.get("period"),
        "matches": matches,
        "breaks": breaks,
        "match_count": len(matches),
        "break_count": len(breaks),
        "net_known_difference": round(total_break, 2),
        "status": "clear" if not breaks else "exceptions_require_review",
        "human_review_required": True,
    }


def screen_kyc(payload: Dict[str, Any]) -> Dict[str, Any]:
    required_docs = set(payload.get("required_documents", ["registration", "beneficial_ownership", "proof_of_address", "tax_form"]))
    provided_docs = set(payload.get("provided_documents", []))
    missing = sorted(required_docs - provided_docs)
    flags = []
    if payload.get("sanctions_hit"):
        flags.append("sanctions_hit")
    if payload.get("pep_hit"):
        flags.append("pep_hit")
    if payload.get("adverse_media"):
        flags.append("adverse_media")
    if payload.get("high_risk_jurisdiction"):
        flags.append("high_risk_jurisdiction")
    if not payload.get("beneficial_owners"):
        flags.append("missing_beneficial_owners")
    if missing:
        flags.append("missing_required_documents")

    if "sanctions_hit" in flags:
        status = "escalation_required"
    elif flags:
        status = "needs_review"
    else:
        status = "complete_for_human_review"

    return {
        "entity_name": payload.get("entity_name"),
        "jurisdiction": payload.get("jurisdiction"),
        "risk_rating": payload.get("risk_rating", "unknown"),
        "missing_documents": missing,
        "flags": flags,
        "status": status,
        "approval": "not_approved_agent_cannot_approve_kyc",
        "human_review_required": True,
    }


def prepare_meeting_brief(payload: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "client_or_company": payload.get("client_or_company"),
        "meeting_date": payload.get("meeting_date"),
        "objective": payload.get("objective"),
        "attendees": payload.get("attendees", []),
        "relationship_context": payload.get("relationship_context", []),
        "recent_developments": payload.get("recent_developments", []),
        "open_items": payload.get("open_items", []),
        "suggested_questions": payload.get("suggested_questions", []),
        "risks_or_sensitivities": payload.get("risks_or_sensitivities", []),
        "next_steps": payload.get("next_steps", []),
        "distribution_note": "Internal briefing draft only. Review before client use.",
        "human_review_required": True,
    }


def review_earnings(payload: Dict[str, Any]) -> Dict[str, Any]:
    reported = payload.get("reported", {})
    consensus = payload.get("consensus", {})
    variances = {}
    for k, v in reported.items():
        if k in consensus:
            variances[k] = round(_num(v) - _num(consensus[k]), 2)
    return {
        "company": payload.get("company"),
        "ticker": payload.get("ticker"),
        "period": payload.get("period"),
        "reported": reported,
        "consensus": consensus,
        "variances": variances,
        "guidance_changes": payload.get("guidance_changes", []),
        "management_commentary": payload.get("management_commentary", []),
        "model_update_flags": payload.get("model_update_flags", []),
        "draft_note_sections": {
            "headline": payload.get("headline", "Earnings review draft"),
            "what_changed": payload.get("what_changed", []),
            "why_it_matters": payload.get("why_it_matters", []),
            "risks": payload.get("risks", []),
            "follow_up_questions": payload.get("follow_up_questions", []),
        },
        "rating_action": "none_agent_does_not_issue_recommendations",
        "human_review_required": True,
    }
